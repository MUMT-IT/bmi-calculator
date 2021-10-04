import math
import openpyxl
from openpyxl import Workbook
import datetime
from tkinter import*
from tkinter import messagebox
from tkinter.messagebox import askyesno

root = Tk()

root.title("BMI calculator")
root.geometry("500x500")
userNameInput = StringVar()
weightInput = DoubleVar()
heightInput = DoubleVar()
userNameLabel = Label(root,text = "ชื่อผู้ใช้")
userNameEntry = Entry(root,width = 30,textvariable=userNameInput)
weightLabel = Label(root,text = "น้ำหนัก(กก.)")
weightEntry = Entry(root,width = 30,textvariable=weightInput)
heightLabel = Label(root,text="ส่วนสูง(ซม.)")
heightEntry = Entry(root,width = 30,textvariable=heightInput)

height1 = heightInput.get()
weight1 = weightInput.get()

userNameLabel.grid(row=0,column=0)
userNameEntry.grid(row=0,column=1)
weightLabel.grid(row=1,column=0)
weightEntry.grid(row=1,column=1)
heightLabel.grid(row=2,column=0)
heightEntry.grid(row=2,column=1)


#------------------------Prepare database-----------------------

filePath = "C:/Users/Peerapat/Desktop/bmi-calculator/Database.xlsx"
wb = openpyxl.load_workbook(filePath)


#---------------------------------------------------------------

#--------------Check for existing name in database--------------
def checkName() :
	sheetNames = wb.sheetnames
	existingName = 0
	name = userNameInput.get()
	if len(name) == 0:
		messagebox.showerror("Error","กรุณากรอกชิ่อผู้ใช้งาน")
		return
	else :	
		for x in range(len(sheetNames)):
			if name == sheetNames[x] :
				existingName +=1

		if existingName > 0 :
			ws = wb[name]
		else :
			confirm = askyesno(title = "Confirmation",message=f"ไม่พบชื่อผู้ใช้งาน {name} ต้องการสร้างรายชื่อใหม่หรือไม่")
			if confirm :
				ws = wb.create_sheet(name)
				ws['A1'] = name
				ws['A2'] = 'เวลา'
				ws['B2'] = 'น้ำหนัก(กก.)'
				ws['C2'] = 'ส่วนสูง(ซม.)'
				ws['D2'] = 'BMI'
				wb.save(filePath)
		return ws
#---------------------------------------------------------------

def newCheckUserName():
	userName = userNameInput.get()
	userNameList = wb.sheetnames
	if userName in userNameList :
		ws = wb[userName]
		return ws
	else:
		messagebox.showerror('Error','ไม่พบชื่อผู้ใช้งาน {userName}')


def createNewUserName():
	userName = userNameInput.get()
	if userName is not None:
			confirm = askyesno(title = "Confirmation",message=f"ต้องการสร้างชื่อผู้ใช้งาน {userName} หรือไม่")
			if confirm :
				ws = wb.create_sheet(userName)
				ws['A1'] = userName
				ws['A2'] = 'เวลา'
				ws['B2'] = 'น้ำหนัก(กก.)'
				ws['C2'] = 'ส่วนสูง(ซม.)'
				ws['D2'] = 'BMI'
				wb.save(filePath)
	else :
		messagebox.showerror("Error","กรุณากรอกชิ่อผู้ใช้งาน")


def BMIresult(BMI):
	if BMI < 18.5 :
		result = "ผอม"
	elif 18.5 <= BMI < 22.9 :
		result = "สุขภาพดี"
	elif 22.9 <= BMI < 24.9 :
		result = "ท้วม"
	elif 24.9 <= BMI < 29.9 :
		result = "อ้วน"
	elif 29.9 <= BMI:
		result = "อ้วนมาก"
	return result

def calculateBMI():
	height = heightInput.get()
	weight = weightInput.get()
	ws = checkName()
	if ws is not None:
		if weight == 0 or weight == None:
			messagebox.showerror("Error","กรุณากรอกข้อมูลน้ำหนัก")
		if weight < 0:
			messagebox.showerror("Error","น้ำหนักไม่ควรติดลบ")
		if type(weight) != float :
			messagebox.showerror("Error","กรุณากรอกข้อมูลน้ำหนักเป็นตัวเลข")
		if height == 0 or weight == None:
			messagebox.showerror("Error","กรุณากรอกข้อมูลความสูง")
		if height < 0:
			messagebox.showerror("Error","ส่วนสูงไม่ควรติดลบ")
		if type(height) != float :
			messagebox.showerror("Error","กรุณากรอกข้อมูลความสูงเป็นตัวเลข")

		if weight >= 0 and type(weight) == float and height >= 0 and type(height) == float:
			height2 = math.pow(height/100, 2)
			BMI = round(weight/height2,2)
			result = BMIresult(BMI)
			resultLabel = Label(text=f"BMI ของคุณคือ {BMI} คุณอยู่ในเกณฑ์ {result}")
			resultLabel.grid(row=4,column=1)
			return BMI
			maxRow = ws.max_row+1
			date = datetime.datetime.now()
			ws.cell(maxRow,1).value = date.strftime('%d/%m/%Y %H:%M')
			ws.cell(maxRow,2).value = weight
			ws.cell(maxRow,3).value = height
			ws.cell(maxRow,4).value = BMI
			wb.save(filePath)

def checkValidInput():
	if weight1 == 0 or weight1 == None:
		messagebox.showerror("Error","กรุณากรอกข้อมูลน้ำหนัก")
	if weight1 < 0:
		messagebox.showerror("Error","น้ำหนักไม่ควรติดลบ")
	if type(weight1) != float :
		messagebox.showerror("Error","กรุณากรอกข้อมูลน้ำหนักเป็นตัวเลข")
	if height1 == 0 or weight1 == None:
		messagebox.showerror("Error","กรุณากรอกข้อมูลความสูง")
	if height1 < 0:
		messagebox.showerror("Error","ส่วนสูงไม่ควรติดลบ")
	if type(height1) != float :
		messagebox.showerror("Error","กรุณากรอกข้อมูลความสูงเป็นตัวเลข")
	if weight1 >= 0 and type(weight1) == float and height1 >= 0 and type(height1) == float:
		return True
		

		
def recordData(BMI):
	weight = weightInput.get()
	height = heightInput.get()
	ws = newCheckUserName()
	if ws is not None :
		maxRow = ws.max_row+1
		date = datetime.datetime.now()
		ws.cell(maxRow,1).value = date.strftime('%d/%m/%Y %H:%M')
		ws.cell(maxRow,2).value = weight
		ws.cell(maxRow,3).value = height
		ws.cell(maxRow,4).value = BMI
		wb.save(filePath)

def showData():
	ws = checkName()
	if ws is not None:
		maxRow = ws.max_row+1
		for row in range(2,maxRow):
			for col in range(1,5):
				dataLabel = Label(root,text=ws.cell(row,col).value)
				dataLabel.grid(row=row+3,column=col-1)

def clearData():
	ws = checkName()
	maxRow = ws.max_row+1
	ws.delete_rows(3,maxRow)
	wb.save(filePath)

def deleteUser():
	ws = checkName()
	name = userNameInput.get()
	if ws is not None:
		confirm = askyesno(title="Confirmation",message = f"ต้องการลบข้อมูลของ {name} ?")
		if confirm :
			wb.remove(ws)
			wb.save(filePath)

def resetForm():
	userNameEntry.delete(0,'end')
	weightEntry.delete(0,'end')
	heightEntry.delete(0,'end')

calBtn = Button(root,width=20,text="คำนวณ BMI",command=calculateBMI)
showBtn = Button(root,width=20,text="เรียกดูข้อมูล",command=showData)
resetBtn = Button(root,width=20,text="รีเซทฟอร์ม",command=resetForm)
clearBtn = Button(root,width=20,text="ลบข้อมูล",command= deleteUser)

calBtn.grid(row=3,column=0)
showBtn.grid(row=3,column=1)
resetBtn.grid(row=3,column=2)
clearBtn.grid(row=3,column=3)

root.mainloop()
