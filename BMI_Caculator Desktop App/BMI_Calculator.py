import math
import openpyxl
from openpyxl import Workbook
import datetime
from tkinter import*
from tkinter import messagebox
from tkinter import ttk
from tkinter.messagebox import ERROR, askyesno

root = Tk()

root.title("BMI calculator")
root.geometry("400x500")

userNameInput = StringVar()
weightInput = DoubleVar()
heightInput = DoubleVar()


userNameLabel = Label(root,text = "ชื่อผู้ใช้งาน")
userNameEntry = Entry(root,width = 20 ,textvariable=userNameInput)
weightLabel = Label(root,text = "น้ำหนัก(กก.)")
weightEntry = Entry(root,width = 20,textvariable=weightInput)
heightLabel = Label(root,text="ส่วนสูง(ซม.)")
heightEntry = Entry(root,width = 20,textvariable=heightInput)

userNameLabel.grid(row=0,column=0)
userNameEntry.grid(row=0,column=1)
weightLabel.grid(row=1,column=0)
weightEntry.grid(row=1,column=1)
heightLabel.grid(row=2,column=0)
heightEntry.grid(row=2,column=1)



columns = ('#1','#2','#3','#4','#5')
dataTreeview = ttk.Treeview(root,columns=columns,show='headings')
dataTreeview.column('#0',width=0,stretch=NO)
dataTreeview.column('#1',width=100,anchor=CENTER)
dataTreeview.column('#2',width=60,anchor=CENTER)
dataTreeview.column('#3',width=60,anchor=CENTER)
dataTreeview.column('#4',width=60,anchor=CENTER)
dataTreeview.column('#5',width=60,anchor=CENTER)

dataTreeview.heading('#1',text='เวลา')
dataTreeview.heading('#2',text='น้ำหนัก')
dataTreeview.heading('#3',text='ส่วนสูง')
dataTreeview.heading('#4',text='BMI')
dataTreeview.heading('#5',text='เกณฑ์')
dataList = []
#------------------------Prepare database-----------------------

filePath = "C:/Users/Peerapat/Desktop/bmi-calculator/BMI_Caculator Desktop App/Database.xlsx"
wb = openpyxl.load_workbook(filePath)

#--------------Check for existing name in database--------------

def checkUserName():
	userName = userNameInput.get()
	userNameList = wb.sheetnames
	if userName != '' :
		if userName in userNameList :
			return True
		else:
			messagebox.showerror('Error',f'ไม่พบชื่อผู้ใช้งาน {userName}')
			return False
	else:
		messagebox.showerror('Error',"กรุณากรอกชิ่อผู้ใช้งาน")

def createNewUserName():
	userName = userNameInput.get()
	if userName != '':
			if  checkUserName():
				messagebox.showerror('Error',f"ชื่อผู้ใช้งาน {userName} มีอยู่แล้ว")
			else:
				confirm = askyesno(title = "Confirmation",message=f"ต้องการสร้างชื่อผู้ใช้งาน {userName} หรือไม่")
				if confirm :
					ws = wb.create_sheet(userName)
					ws['A1'] = userName
					ws['A2'] = 'เวลา'
					ws['B2'] = 'น้ำหนัก(กก.)'
					ws['C2'] = 'ส่วนสูง(ซม.)'
					ws['D2'] = 'BMI'
					ws['E2'] = 'เกณฑ์'
					wb.save(filePath)
	else :
		messagebox.showerror("Error","กรุณากรอกชิ่อผู้ใช้งาน")

def validateInput():
	try:
		weight = weightInput.get()
	except TclError:
		messagebox.showerror("Error","กรุณากรอกน้ำหนักเป็นตัวเลข")
	try:
		height = heightInput.get()
	except TclError:
		messagebox.showerror("Error","กรุณากรอกส่วนสูงเป็นตัวเลข")
	if weight == 0 or weight is None or weight == '':
		messagebox.showerror("Error","กรุณากรอกข้อมูลน้ำหนัก")
		return False
	if height == 0 or height is None or height == '':
		messagebox.showerror("Error","กรุณากรอกข้อมูลส่วนสูง")
		return False
	if type(weight) is float and type(height) is float and weight > 0 and height > 0:
		return True

def BMIresult(BMI):
	if BMI < 18.5 :
		result = "ผอม"
	elif 18.5 <= BMI < 22.9 :
		result = "สุขภาพดี"
	elif 22.9 <= BMI < 24.9 :
		result = "ท้วม"
	elif 24.9 <= BMI < 29.9 :
		result = "อ้วน"
	elif 29.9 <= BMI:
		result = "อ้วนมาก"
	return result

def calculateBMI():
	if validateInput():
		height = heightInput.get()
		weight = weightInput.get()
		height2 = math.pow(float(height)/100, 2)
		BMI = round(float(weight)/height2,2)
		result = BMIresult(BMI)
		resultLabel = Label(root,width = 20,text=f"BMI {BMI} {result}")
		resultLabel.grid(row=6,column=1)
		return BMI

def recordData():
	if checkUserName():
		if validateInput():
			userName = userNameInput.get()
			height = heightInput.get()
			weight = weightInput.get()
			ws = wb[userName]
			BMI = calculateBMI()
			result = BMIresult(BMI)
			maxRow = ws.max_row+1
			date = datetime.datetime.now()
			ws.cell(maxRow,1).value = date.strftime('%d/%m/%Y %H:%M')
			ws.cell(maxRow,2).value = weight
			ws.cell(maxRow,3).value = height
			ws.cell(maxRow,4).value = BMI
			ws.cell(maxRow,5).value = result
			wb.save(filePath)

def showData():
	if len(dataList) != 0:
		for data in dataTreeview.get_children():
			dataTreeview.delete(data)
		dataList.clear()
	userName = userNameInput.get()
	if checkUserName():
		ws = wb[userName]
		maxRow = ws.max_row+1
		for row in range(3,maxRow):
			dataList.append((ws.cell(row,1).value,ws.cell(row,2).value,ws.cell(row,3).value,ws.cell(row,4).value,ws.cell(row,5).value))
		for data in dataList:
			dataTreeview.insert('',END,values=data)	
		dataTreeview.grid(row=7,column=1)
		

def deleteUser():
	userName = userNameInput.get()
	if checkUserName():
		ws = wb[userName]
		confirm = askyesno(title="Confirmation",message = f"ต้องการลบข้อมูลของ {userName} ?")
		if confirm :
			wb.remove(ws)
			wb.save(filePath)
			dataTreeview.destroy()

def resetForm():
	userNameEntry.delete(0,'end')
	weightEntry.delete(0,'end')
	heightEntry.delete(0,'end')
	dataTreeview.destroy()

calBtn = Button(root,width=20,text="คำนวณ BMI",command=calculateBMI)
newUserBtn = Button(root,width=20,text="สร้างชื่อผู้ใช้ใหม่",command=createNewUserName)
recBtn = Button(root,width=20,text="บันทึกข้อมูล",command=recordData)
showBtn = Button(root,width=20,text="เรียกดูข้อมูล",command=showData)
delBtn = Button(root,width=20,text="ลบข้อมูล",command= deleteUser)
resetBtn = Button(root,width=20,text="รีเซทฟอร์ม",command=resetForm)

calBtn.grid(row=3,column=0)
newUserBtn.grid(row=3,column=1)
recBtn.grid(row=4,column=0)
showBtn.grid(row=4,column=1)
resetBtn.grid(row=5,column=0)
delBtn.grid(row=5,column=1)

root.mainloop()
